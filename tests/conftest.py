import csv
import datetime
import tempfile

import pytest
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from pytz import timezone

from import_me.columns import Column
from import_me.parsers.base import BaseParser
from import_me.parsers.xlsx import BaseXLSXParser


choices_classifier_no_processor = [
    ['e', 'A'],
    ['d', lambda x: x in ['test', 'test2']],
    ['a', lambda x: 0 <= x <= 10],
    ['b', lambda x: 10 <= x <= 100],
    ['c', lambda x: isinstance(x, str)],
]


choices_classifier_integer_processor = [
    ['a', lambda x: 0 <= x <= 10],
    ['b', lambda x: 10 <= x <= 100],
]


choices_classifier_datetime_processor = [
    ['a', lambda x: datetime.date(2020, 1, 1) <= x <= datetime.date(2020, 12, 31)],
    ['b', lambda x: datetime.date(2021, 1, 1) <= x <= datetime.date(2021, 12, 31)],
]


datetime_for_test = datetime.datetime(2019, 7, 20, 12, 43, 52)
user_timezone = timezone('Europe/Moscow')
datetime_for_test_with_user_timezone = datetime_for_test.astimezone(user_timezone)


@pytest.fixture
def base_parser():
    class Parser(BaseParser):
        add_file_path = True
        add_row_index = True
        skip_empty_rows = True
        columns = [Column('first_name', index=1)]

    return Parser(file_path='test_file_path')


@pytest.fixture
def cell_factory():
    def cell(value):
        return value
    return cell


@pytest.fixture
def row_factory():
    def row(values):
        return list(values)

    return row


@pytest.fixture
def xlsx_parser():
    class Parser(BaseXLSXParser):
        add_file_path = True
        add_row_index = True
        skip_empty_rows = True
        columns = [Column('first_name', index=1)]

    return Parser(file_path='test_file_path')


@pytest.fixture
def xlsx_cell_factory():
    def cell(value):
        return Cell(Worksheet(Workbook()), value=value)

    return cell


@pytest.fixture
def xlsx_row_factory(xlsx_cell_factory):
    def row(values):
        return [xlsx_cell_factory(value) for value in values]

    return row


@pytest.fixture
def workbook_factory():
    def _workbook_factory(
        header=None, data=None, header_row_index=0, data_row_index=1, worksheets_count=1,
    ):
        wb = Workbook()
        wb.remove(wb.active)  # remove default worksheet for clear creation

        for worksheet_index in range(worksheets_count):
            ws = wb.create_sheet(title=str(worksheet_index), index=worksheet_index)
            if header is not None:
                for _row_index in range(header_row_index):
                    ws.append([''] * len(header))
                ws.append(header)

            if data is not None:
                for _row_index in range(header_row_index, data_row_index - 1):
                    ws.append([''] * len(data))
                for row in data:
                    ws.append(row)
        return wb
    return _workbook_factory


@pytest.fixture
def xlsx_file_factory(workbook_factory):
    def _xlsx_file_factory(
        header=None, data=None, header_row_index=0, data_row_index=1, worksheets_count=1,
    ):
        xlsx_file = tempfile.NamedTemporaryFile(suffix='.xlsx')
        wb = workbook_factory(header, data, header_row_index, data_row_index, worksheets_count)
        wb.save(xlsx_file.name)
        return xlsx_file
    return _xlsx_file_factory


@pytest.fixture
def csv_file_factory():
    def _csv_file_factory(
        header=None, data=None, header_row_index=0, data_row_index=1, file_kwargs=None, writer_kwargs=None,
    ):
        file_kwargs = file_kwargs or {}
        writer_kwargs = writer_kwargs or {}
        csv_file = tempfile.NamedTemporaryFile(suffix='.csv')
        with open(csv_file.name, 'w', **file_kwargs) as file:
            writer = csv.writer(file, **writer_kwargs)

            if header is not None:
                for _row_index in range(header_row_index):
                    writer.writerow([''] * len(header))
                writer.writerow(header)
            if data is not None:
                for _row_index in range(header_row_index, data_row_index - 1):
                    writer.writerow([''] * len(data))
                for row in data:
                    writer.writerow(row)
        return csv_file
    return _csv_file_factory


def raise_(ex, *args, **kwargs):
    raise ex(*args, **kwargs)
