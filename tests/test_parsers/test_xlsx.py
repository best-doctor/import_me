import sys
from unittest.mock import MagicMock

import pytest
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.workbook import Workbook

from import_me.columns import Column
from import_me.exceptions import StopParsing
from import_me.parsers.xlsx import BaseXLSXParser, BaseMultipleSheetsXLSXParser
from import_me.processors import FloatProcessor

DEFAULT_WORKBOOK_DATA = {
    'header': ['First Name', 'Last Name'],
    'data': [['Ivan', 'Ivanov'], ['Petr', 'Petrov']],
}

DEFAULT_PARSER_COLUMNS = [
    Column('first_name', index=0, header='First Name'),
    Column('last_name', index=1, header='Last Name'),
]


def test_base_xlsx_parser(xlsx_file_factory):
    class XLSXParser(BaseXLSXParser):
        columns = DEFAULT_PARSER_COLUMNS

    xlsx_file = xlsx_file_factory(**DEFAULT_WORKBOOK_DATA)
    parser = XLSXParser(file_path=xlsx_file.name)

    parser()

    assert parser.has_errors is False
    assert parser.cleaned_data == [
        {
            'first_name': 'Ivan',
            'last_name': 'Ivanov',
            'row_index': 1,
        },
        {
            'first_name': 'Petr',
            'last_name': 'Petrov',
            'row_index': 2,
        },
    ]


def test_base_xlsx_parser_accepts_file_object(xlsx_file_factory):
    class XLSXParser(BaseXLSXParser):
        columns = DEFAULT_PARSER_COLUMNS

    xlsx_file = xlsx_file_factory(**DEFAULT_WORKBOOK_DATA)
    parser = XLSXParser(file_contents=xlsx_file)

    parser()

    assert parser.has_errors is False
    assert parser.cleaned_data == [
        {
            'first_name': 'Ivan',
            'last_name': 'Ivanov',
            'row_index': 1,
        },
        {
            'first_name': 'Petr',
            'last_name': 'Petrov',
            'row_index': 2,
        },
    ]


@pytest.mark.xfail(
    sys.version_info >= (3, 9),
    reason="Xlrd doesn't work with xlsx on python 3.9",
)
def test_base_xlsx_parser_accepts_file_content(xlsx_file_factory):
    class XLSXParser(BaseXLSXParser):
        columns = DEFAULT_PARSER_COLUMNS

    xlsx_file = xlsx_file_factory(**DEFAULT_WORKBOOK_DATA)
    parser = XLSXParser(file_contents=xlsx_file.read())
    parser._load_workbook_from_xlsx = MagicMock(side_effect=InvalidFileException)

    parser()

    assert parser.has_errors is False
    assert parser.cleaned_data == [
        {
            'first_name': 'Ivan',
            'last_name': 'Ivanov',
            'row_index': 1,
        },
        {
            'first_name': 'Petr',
            'last_name': 'Petrov',
            'row_index': 2,
        },
    ]


def test_base_xlsx_parser_without_header(xlsx_file_factory):
    class XLSXParser(BaseXLSXParser):
        columns = [
            Column('first_name', index=0),
            Column('last_name', index=1),
        ]
        first_data_row_index = 0

    xlsx_file = xlsx_file_factory(
        data=[
            ['Ivan', 'Ivanov'],
            ['Petr', 'Petrov'],
        ],
        data_row_index=0,
    )
    parser = XLSXParser(file_path=xlsx_file.name)

    parser()

    assert parser.has_errors is False
    assert parser.cleaned_data == [
        {
            'first_name': 'Ivan',
            'last_name': 'Ivanov',
            'row_index': 0,
        },
        {
            'first_name': 'Petr',
            'last_name': 'Petrov',
            'row_index': 1,
        },
    ]


def test_base_xlsx_parser_clean_column():
    class Parser(BaseXLSXParser):
        columns = [
            Column('last_name', index=0),
            Column('first_name', index=1),
        ]

        def clean_column_last_name(self, value):
            return f'Modified {value}'

    parser = Parser(file_path=None)

    assert parser.clean_column(parser.columns[0], 'Test Last Name') == 'Modified Test Last Name'
    assert parser.clean_column(parser.columns[1], 'Test First Name') == 'Test First Name'


@pytest.mark.parametrize(
    'header_row_index, first_data_row_index, expected_result',
    (
        (None, None, None),
        (10, None, 10),
        (None, 20, 19),
    ),
)
def test_base_xlsx_parser_header_row_offset(header_row_index, first_data_row_index, expected_result):
    parser = BaseXLSXParser(file_path=None)
    parser.header_row_index = header_row_index
    parser.first_data_row_index = first_data_row_index

    assert parser.header_row_offset == expected_result


@pytest.mark.parametrize(
    'header_row_idx',
    (0, 1, 2),
)
def test_validate_worksheet_headers(header_row_idx, workbook_factory):
    class Parser(BaseXLSXParser):
        header_row_index = header_row_idx
        columns = [
            Column('column1', index=0, header='колонка1'),
            Column('column2', index=1, header='колонка2', validate_header=False),
            Column('column3', index=2),
            Column('column4', index=3, header='колонка4', validate_header=True),
            Column('column5', index=4, header='Колонка5'),
            Column('column6', index=5, header='колонка6'),
        ]

    parser = Parser()
    workbook = workbook_factory(
        header=['колонка1', 'test', 'test', 'колонка4', '   колонка5   ', 'КоЛоНка6'],
        data=[],
        header_row_index=header_row_idx,
    )

    parser.validate_worksheet_headers(worksheet=workbook.active)


@pytest.mark.parametrize(
    'column_header, column_index, wb_headers',
    (
        ('колонка1', 0, ['не колонка1']),
        ('колонка1', 1, ['колонка1']),
        ('колонка1', 0, ['', 'колонка1']),
        ('колонка1', 0, [None]),
    ),
)
def test_validate_worksheet_headers_errors(
    column_header, column_index, wb_headers, workbook_factory,
):
    class Parser(BaseXLSXParser):
        columns = [
            Column('column1', index=column_index, header=column_header),
        ]

    parser = Parser()
    workbook = workbook_factory(header=wb_headers, data=[])

    with pytest.raises(StopParsing):
        parser.validate_worksheet_headers(worksheet=workbook.active)


@pytest.mark.parametrize(
    'parser_skip_empty_rows, file_data, expected_data',
    (
        (True, [[None], ['column1_data']], [{'column1': 'column1_data'}]),
        (False, [[None], ['column1_data']], [{'column1': None}, {'column1': 'column1_data'}]),
    ),
)
def test_parser_skip_empty_row(parser_skip_empty_rows, file_data, expected_data, xlsx_file_factory):
    class Parser(BaseXLSXParser):
        skip_empty_rows = parser_skip_empty_rows
        add_file_path = False
        add_row_index = False
        columns = [
            Column('column1', index=0),
        ]

    xlsx_file = xlsx_file_factory(header=['column1'], data=file_data)
    parser = Parser(file_path=xlsx_file.file)

    parser()

    assert parser.has_errors is False
    assert parser.cleaned_data == expected_data


def test_parser_unique_column(xlsx_file_factory):
    class Parser(BaseXLSXParser):
        columns = [
            Column('id', index=0, unique=True),
        ]

    xlsx_file = xlsx_file_factory(
        header=['id'],
        data=[
            ['1'],
            ['2'],
            ['1'],
        ],
    )
    parser = Parser(file_path=xlsx_file.file)

    parser()

    assert parser.has_errors is True
    assert parser.errors == ['row: 3, column: 0, value 1 is a duplicate of row 1']
    assert parser.cleaned_data == [
        {'id': '1', 'row_index': 1},
        {'id': '2', 'row_index': 2},
    ]


def test_parser_unique_column_without_header(xlsx_file_factory):
    class Parser(BaseXLSXParser):
        first_data_row_index = 0
        columns = [
            Column('id', index=0, unique=True),
        ]

    xlsx_file = xlsx_file_factory(
        data_row_index=0,
        data=[
            ['1'],
            ['2'],
            ['1'],
        ],
    )
    parser = Parser(file_path=xlsx_file.file)

    parser()

    assert parser.has_errors is True
    assert parser.errors == ['row: 2, column: 0, value 1 is a duplicate of row 0']
    assert parser.cleaned_data == [
        {'id': '1', 'row_index': 0},
        {'id': '2', 'row_index': 1},
    ]


def test_parser_unique_together_column(xlsx_file_factory):
    class Parser(BaseXLSXParser):
        columns = [
            Column('first_name', index=0),
            Column('last_name', index=1),
            Column('middle_name', index=2),
        ]
        unique_together = [
            ['first_name', 'last_name'],
            ['last_name', 'middle_name'],
        ]

    xlsx_file = xlsx_file_factory(
        header=['first_name', 'last_name', 'middle_name'],
        data=[
            ['Ivan', 'Ivanov', 'Ivanovich'],
            ['Ivan', 'Ivanov', 'Petrovich'],
            ['Petr', 'Ivanov', 'Ivanovich'],
            ['Petr', 'Petrov', 'Petrovich'],
        ],
    )
    parser = Parser(file_path=xlsx_file.file)

    parser()

    assert parser.has_errors is True
    assert parser.errors == [
        'row: 2, first_name (Ivan), last_name (Ivanov) is a duplicate of row 1',
        'row: 3, last_name (Ivanov), middle_name (Ivanovich) is a duplicate of row 1',
    ]
    assert parser.cleaned_data == [
        {'first_name': 'Ivan', 'last_name': 'Ivanov', 'middle_name': 'Ivanovich', 'row_index': 1},
        {'first_name': 'Petr', 'last_name': 'Petrov', 'middle_name': 'Petrovich', 'row_index': 4},
    ]


@pytest.mark.parametrize(
    'header_row_index, first_data_row_index, expected_result',
    (
        (None, None, None),
        (10, None, 10),
        (None, 20, 19),
    ),
)
def test_base_multiple_sheets_xlsx_parser_header_row_offset(
    header_row_index, first_data_row_index, expected_result,
):
    parser = BaseMultipleSheetsXLSXParser(file_path=None)
    parser.header_row_index = header_row_index
    parser.first_data_row_index = first_data_row_index

    assert parser.header_row_offset == expected_result


@pytest.mark.parametrize(
    'column_header, column_index, wb_headers',
    (
        ('колонка1', 0, ['не колонка1']),
        ('колонка1', 1, ['колонка1']),
        ('колонка1', 0, ['', 'колонка1']),
        ('колонка1', 0, [None]),
    ),
)
def test_base_multiple_sheets_xlsx_parser_validate_worksheet_headers_errors(
    column_header, column_index, wb_headers, workbook_factory,
):
    class Parser(BaseMultipleSheetsXLSXParser):
        columns = [
            Column('column1', index=column_index, header=column_header),
        ]
    parser = Parser()
    workbook = workbook_factory(header=wb_headers, data=[])

    assert parser._validate_worksheet_headers(worksheet=workbook.active) is not None


@pytest.mark.parametrize(
    ('headers_error', 'title_error', 'errors_len'),
    (
        (None, None, 0),
        ('error', 'error', 2),
        ('error', None, 1),
        (None, 'error', 1),
    ),
    ids=['no-errors', 'all-errors', 'header-error', 'title-error'],
)
def test_base_multiple_sheets_xlsx_parser_validate_worksheet(
    headers_error, title_error, errors_len, workbook_factory,
):
    class Parser(BaseMultipleSheetsXLSXParser):
        columns = [
            Column('column1', index=0),
        ]
    parser = Parser()
    parser._validate_worksheet_headers = MagicMock(return_value=headers_error)
    parser._validate_worksheet_title = MagicMock(return_value=title_error)
    workbook = workbook_factory()

    valid = parser._validate_worksheet(workbook.active)

    assert valid is not bool(errors_len)
    assert len(parser.errors) == errors_len
    parser._validate_worksheet_headers.assert_called_once()
    parser._validate_worksheet_title.assert_called_once()


@pytest.mark.parametrize('read_only', [True, False], ids=['ro_true', 'ro_false'])
def test_base_multiple_sheets_xlsx_parser_load_workbook(xlsx_file_factory, read_only):
    class MultipleSheetsXLSXParser(BaseMultipleSheetsXLSXParser):
        read_only_workbook = read_only
        columns = DEFAULT_PARSER_COLUMNS

    xlsx_file = xlsx_file_factory(**DEFAULT_WORKBOOK_DATA)
    parser = MultipleSheetsXLSXParser(file_path=xlsx_file.name)
    workbook = parser._load_workbook()

    assert isinstance(workbook, Workbook)
    assert workbook.read_only == read_only
    assert workbook.data_only == read_only


def test_base_multiple_sheets_xlsx_parser_iterate_worksheet_rows(workbook_factory):
    workbook = workbook_factory(**DEFAULT_WORKBOOK_DATA)
    parser = BaseMultipleSheetsXLSXParser(file_path=None)
    worksheet_iterator = parser._iterate_worksheet_rows(workbook.active)

    row_index, row = worksheet_iterator.__next__()

    assert row_index == 1
    assert row == DEFAULT_WORKBOOK_DATA['data'][0]


def test_base_multiple_sheets_xlsx_parser_parse_worksheet(workbook_factory):
    workbook = workbook_factory(**DEFAULT_WORKBOOK_DATA)

    class MultipleSheetsXLSXParser(BaseMultipleSheetsXLSXParser):
        columns = DEFAULT_PARSER_COLUMNS
    parser = MultipleSheetsXLSXParser(file_path=None)

    parser._parse_worksheet(workbook.active)

    assert parser.cleaned_data == [
        {'worksheet': '0', 'first_name': 'Ivan', 'last_name': 'Ivanov', 'row_index': 1},
        {'worksheet': '0', 'first_name': 'Petr', 'last_name': 'Petrov', 'row_index': 2},
    ]


def test_base_multiple_sheets_xlsx_parser_parse_worksheet_with_errors(workbook_factory):
    workbook = workbook_factory(**DEFAULT_WORKBOOK_DATA)

    class MultipleSheetsXLSXParser(BaseMultipleSheetsXLSXParser):
        columns = [
            Column('first_name', index=0, header='First Name', processor=FloatProcessor()),
            Column('last_name', index=1, header='Last Name', processor=FloatProcessor()),
        ]
    parser = MultipleSheetsXLSXParser(file_path=None)

    parser._parse_worksheet(workbook.active)

    assert parser.cleaned_data == []
    assert parser.errors == [
        'worksheet: 0, row: 1, column: 0, Ivan is not a floating point number.',
        'worksheet: 0, row: 1, column: 1, Ivanov is not a floating point number.',
        'worksheet: 0, row: 2, column: 0, Petr is not a floating point number.',
        'worksheet: 0, row: 2, column: 1, Petrov is not a floating point number.',
    ]


def test_base_multiple_sheets_xlsx_parser_parse(xlsx_file_factory):

    class MultipleSheetsXLSXParser(BaseMultipleSheetsXLSXParser):
        columns = DEFAULT_PARSER_COLUMNS

    xlsx_file = xlsx_file_factory(worksheets_count=2, **DEFAULT_WORKBOOK_DATA)
    parser = MultipleSheetsXLSXParser(file_path=xlsx_file.name)

    parser._parse()

    assert parser.cleaned_data == [
        {'worksheet': '0', 'first_name': 'Ivan', 'last_name': 'Ivanov', 'row_index': 1},
        {'worksheet': '0', 'first_name': 'Petr', 'last_name': 'Petrov', 'row_index': 2},
        {'worksheet': '1', 'first_name': 'Ivan', 'last_name': 'Ivanov', 'row_index': 1},
        {'worksheet': '1', 'first_name': 'Petr', 'last_name': 'Petrov', 'row_index': 2},
    ]


def test_base_multiple_sheets_xlsx_parser_parse_data_error(xlsx_file_factory):
    class MultipleSheetsXLSXParser(BaseMultipleSheetsXLSXParser):
        columns = DEFAULT_PARSER_COLUMNS

    xlsx_file = xlsx_file_factory(worksheets_count=2, **DEFAULT_WORKBOOK_DATA)
    parser = MultipleSheetsXLSXParser(file_path=xlsx_file.name)
    parser._parse = MagicMock(side_effect=Exception('Test exception'))

    parser.parse_data()

    assert parser.errors[0] == 'Test exception'
