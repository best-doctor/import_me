from __future__ import annotations
import datetime
import pathlib

import xlrd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet

from bdata_parser.exceptions import ColumnError, ParserError, SkipRow, StopParsing

if False:  # TYPE_CHECKING
    from typing import List, Callable, Dict, Tuple, Any, Type, Optional, Union, IO


class Column:
    def __init__(
        self, name: str, index: int, processor: Callable = None,
        header: str = None, validate_header: bool = True, required: bool = False,
    ):
        self.name = name
        self.index = index
        if processor:
            self.processor = processor
        else:
            self.processor = lambda x: x
        self.header = header
        self.validate_header = validate_header
        self.required = required


class ParserMixin:
    skip_empty_rows: bool = True
    add_file_path: bool = True
    add_row_index: bool = True

    def __init__(self, *args: Any, **kwargs: Any) -> None:
        self.cleaned_data: List[Dict[str, Any]] = []
        self.errors: List[str] = []

    @property
    def has_errors(self) -> bool:
        return bool(self.errors)

    def __call__(self, raise_errors: bool = False, *args: Any, **kwargs: Any) -> None:
        raise NotImplementedError


class BaseParser(ParserMixin):
    columns: List[Column]
    ws_index: int = 0
    header_row_index: Optional[int] = None
    first_data_row_index: int = 2  # 1-based index
    last_data_row_index: Optional[int] = None

    def __init__(self, file_path: pathlib.Path = None, file_contents: IO = None, *args: Any, **kwargs: Any) -> None:
        super().__init__(*args, **kwargs)
        self.file_path = file_path
        self.file_contents = file_contents

    @property
    def header_row_offset(self) -> Optional[int]:
        index = None
        if self.header_row_index is not None:
            index = self.header_row_index
        elif self.first_data_row_index:
            index = self.first_data_row_index - 1
        return index

    def parse(self) -> None:
        data = []

        wb = self.load_workbook()
        self.validate_workbook(wb)

        ws = wb[wb.sheetnames[self.ws_index]]
        self.validate_worksheet(ws)

        row_index = self.first_data_row_index
        for row in ws.iter_rows(min_row=self.first_data_row_index, max_row=self.last_data_row_index):
            try:
                row_data = self.parse_row(row, row_index)
            except SkipRow:
                pass
            else:
                data.append(row_data)

            row_index += 1

        self.cleaned_data = self.clean(data)

    def parse_row(self, row: Tuple, row_index: int) -> Dict:
        row_data = {}
        row_has_errors = False

        for column in self.columns:
            try:
                row_data[column.name] = self.parse_column(row, column, row_index)
            except ColumnError as e:
                row_has_errors = True
                self.add_errors(e.messages, row_index=row_index, col_index=column.index)

        if row_has_errors:
            raise SkipRow('Не обработана т.к. содержит ошибки')

        return self.clean_row(row_data, row, row_index)

    def parse_column(self, row: Tuple, column: Column, row_index: int) -> Any:
        try:
            value = row[column.index].value
        except IndexError:
            value = None

        try:
            value = column.processor(value)
            value = self.clean_column(column, value)
        except StopParsing as e:
            raise e
        except Exception as e:
            raise ColumnError(getattr(e, 'messages', str(e))) from e

        return value

    def clean_row(self, row_data: Dict, row: Tuple, row_index: int) -> Dict:
        if self.skip_empty_rows and all((row_data.get(column.name) is None for column in self.columns)):
            raise SkipRow

        row_data = self.clean_row_required_columns(row_data, row, row_index)

        if self.add_file_path:
            row_data['file_path'] = self.file_path
        if self.add_row_index:
            row_data['row_index'] = row_index

        return row_data

    def clean_row_required_columns(self, row_data: Dict, row: Tuple, row_index: int) -> Dict:
        has_empty_required_columns = False

        for column in self.columns:
            if column.required and row_data.get(column.name) is None:
                self.add_errors(
                    f'Колонка {column.header or column.name} обязательна к заполнению.',
                    row_index=row_index, col_index=column.index,
                )
                has_empty_required_columns = True

        if has_empty_required_columns:
            raise SkipRow(f'В строке {row_index} есть незаполненные колонки.')

        return row_data

    def clean_column(self, column: Column, value: Any) -> Any:
        column_clean_func = getattr(self, f'clean_column_{column.name}', None)
        if column_clean_func:
            value = column_clean_func(value)
        return value

    def clean(self, data: List) -> List:
        return data

    def load_workbook(self) -> Workbook:
        """Загрузка openpyxl Workbook из файла.

        Пробуем загрузить сначала через openpyxl,
        если он не умеет работать с данным типом файлов,
        то читаем файл с помощью xlrd.

        """
        try:
            wb = self._load_workbook_from_xlsx()
        except InvalidFileException:
            wb = self._load_workbook_from_xls()
        return wb

    def validate_workbook(self, workbook: Workbook) -> None:
        pass

    def validate_worksheet(self, worksheet: Worksheet) -> None:
        self.validate_worksheet_headers(worksheet)

    def validate_worksheet_headers(self, worksheet: Worksheet) -> None:
        expected_headers = {
            column.index: column.header.lower()
            for column in self.columns
            if column.header and column.validate_header
        }
        if expected_headers and self.header_row_offset is not None:
            for row in worksheet.iter_rows(min_row=self.header_row_offset):
                columns = {
                    idx: col.value.strip().lower()
                    if isinstance(col.value, str) else col.value
                    for idx, col in enumerate(row) if idx in expected_headers
                }

                if columns != expected_headers:
                    file_path = self.file_path or 'file'
                    raise StopParsing((
                        f'Не верные названия колонок в файле {file_path}. '
                        f'Колонки в файле: {columns}. '
                        f'Ожидаемые колонки: {expected_headers}.'))
                break

    def add_errors(self, messages: Union[str, List], row_index: int = None, col_index: int = None) -> None:
        if not isinstance(messages, list):
            messages = [messages]
        for message in messages:
            error = []
            if row_index is not None:
                error.append(f'строка: {row_index}')
            if col_index is not None:
                error.append(f'колонка: {col_index}')
            error.append(message)
            self.errors.append(', '.join(error))

    def __call__(self, raise_errors: bool = False, *args: Any, **kwargs: Any) -> None:
        try:
            self.parse()
        except Exception as e:
            messages = getattr(e, 'messages', str(e))
            self.add_errors(messages)

        if raise_errors and self.has_errors:
            raise ParserError(self.errors)

    def _load_workbook_from_xlsx(self) -> Workbook:
        return load_workbook(filename=self.file_path or self.file_contents, read_only=True, data_only=True)

    def _load_workbook_from_xls(self) -> Workbook:
        xls_workbook = xlrd.open_workbook(filename=self.file_path, file_contents=self.file_contents)
        xls_sheet = xls_workbook.sheet_by_index(0)
        nrows = xls_sheet.nrows
        ncols = xls_sheet.ncols

        wb = Workbook()
        ws = wb[wb.sheetnames[0]]

        for row in range(nrows):
            for col in range(ncols):
                cell = xls_sheet.cell(row, col)
                value = cell.value
                if value and cell.ctype == 3:
                    value = datetime.datetime(*xlrd.xldate_as_tuple(value, xls_workbook.datemode))
                ws.cell(row=row + 1, column=col + 1).value = value

        return wb


class BaseMultipleFileParser(ParserMixin):
    parser_class: Type[BaseParser]
    dir_path: pathlib.Path
    filename_patterns: List[str] = ['*.xls', '*.xlsx']

    def __init__(self, dir_path: pathlib.Path = None, *args: Any, **kwargs: Any) -> None:
        super().__init__(*args, **kwargs)
        if dir_path:
            self.dir_path = dir_path

    def get_file_paths(self) -> List[pathlib.Path]:
        paths = []
        for filename_pattern in self.filename_patterns:
            for file_path in self.dir_path.glob(filename_pattern):
                paths.append(file_path)
        return sorted(paths)

    def add_errors(self, messages: Union[List, str], file_path: pathlib.Path) -> None:
        if not isinstance(messages, list):
            messages = [messages]
        for message in messages:
            self.errors.append(f'{file_path}, {message}')

    def __call__(self, raise_errors: bool = False, *args: Any, **kwargs: Any) -> None:
        for file_path in self.get_file_paths():
            try:
                parser = self.parser_class(file_path)
                parser(raise_errors=raise_errors)
            except Exception as e:
                messages = getattr(e, 'messages', str(e))
                self.add_errors(messages, file_path)
            else:
                if parser.has_errors:
                    self.add_errors(parser.errors, file_path)
                self.cleaned_data.extend(parser.cleaned_data)

        if raise_errors and self.has_errors:
            raise ParserError(self.errors)
