from __future__ import annotations

import datetime
from typing import TYPE_CHECKING

import xlrd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet

from import_me.parsers.base import BaseParser, ParserMixin
from import_me.exceptions import StopParsing, SkipRow, ParserError

if TYPE_CHECKING:
    from typing import Optional, Iterator, Tuple, Any, List


class BaseXLSXParser(BaseParser):
    ws_index: int = 0
    header_row_index: Optional[int] = None
    first_data_row_index: int = 1
    last_data_row_index: Optional[int] = None

    @property
    def header_row_offset(self) -> Optional[int]:
        index = None
        if self.header_row_index is not None:
            index = self.header_row_index
        elif self.first_data_row_index:
            index = self.first_data_row_index - 1
        return index

    def iterate_file_rows(self) -> Iterator[Tuple[int, List[Any]]]:
        wb = self.load_workbook()
        self.validate_workbook(wb)

        ws = wb[wb.sheetnames[self.ws_index]]
        self.validate_worksheet(ws)

        row_index = self.first_data_row_index
        min_row = self.first_data_row_index + 1 if self.first_data_row_index is not None else None
        max_row = self.last_data_row_index + 1 if self.last_data_row_index is not None else None
        for row in ws.iter_rows(min_row=min_row, max_row=max_row):
            yield row_index, [cell.value for cell in row]
            row_index += 1

    def load_workbook(self) -> Workbook:
        """Загрузка openpyxl Workbook из файла.

        Пробуем загрузить сначала через openpyxl,
        если он не умеет работать с данным типом файлов,
        то читаем файл с помощью xlrd.

        """
        try:
            wb = self._load_workbook_from_xlsx()
        except InvalidFileException:
            wb = self._load_workbook_from_xls()
        return wb

    def validate_workbook(self, workbook: Workbook) -> None:
        pass

    def validate_worksheet(self, worksheet: Worksheet) -> None:
        self.validate_worksheet_headers(worksheet)

    def validate_worksheet_headers(self, worksheet: Worksheet) -> None:
        expected_headers = {
            column.index: column.header.lower()
            for column in self.columns
            if column.header and column.validate_header
        }
        if expected_headers and self.header_row_offset is not None:
            for row in worksheet.iter_rows(min_row=self.header_row_offset + 1):
                columns = {
                    idx: col.value.strip().lower()
                    if isinstance(col.value, str) else col.value
                    for idx, col in enumerate(row) if idx in expected_headers
                }

                if columns != expected_headers:
                    file_path = self.file_path or 'file'
                    raise StopParsing((
                        f'Incorrect column names in the file: {file_path}. '
                        f'Columns in file: {columns}. '
                        f'Expected columns: {expected_headers}.'))
                break

    def _load_workbook_from_xlsx(self) -> Workbook:
        return load_workbook(filename=self.file_path or self.file_contents, read_only=True, data_only=True)

    def _load_workbook_from_xls(self) -> Workbook:
        file_contents = self.file_contents
        try:
            file_contents = self.file_contents.read()  # type: ignore
        except AttributeError:
            # This is likely the case when someone passed an
            # actual file content instead of just opened file
            pass

        xls_workbook = xlrd.open_workbook(filename=self.file_path, file_contents=file_contents)
        xls_sheet = xls_workbook.sheet_by_index(0)
        nrows = xls_sheet.nrows
        ncols = xls_sheet.ncols

        wb = Workbook()
        ws = wb[wb.sheetnames[0]]

        for row in range(nrows):
            for col in range(ncols):
                cell = xls_sheet.cell(row, col)
                value = cell.value
                if value and cell.ctype == 3:
                    value = datetime.datetime(*xlrd.xldate_as_tuple(value, xls_workbook.datemode))
                ws.cell(row=row + 1, column=col + 1).value = value

        return wb


class BaseMultipleXLSXFileParser(ParserMixin):
    filename_patterns: List[str] = ['*.xls', '*.xlsx']


class BaseMultipleSheetsXLSXParser(BaseParser):
    header_row_index: Optional[int] = None
    first_data_row_index: int = 1
    last_data_row_index: Optional[int] = None
    read_only_workbook: bool = True

    def __init__(self, *args: Any, **kwargs: Any) -> None:
        super().__init__(*args, **kwargs)
        self.workbook: Workbook = None

    @property
    def header_row_offset(self) -> Optional[int]:
        index = None
        if self.header_row_index is not None:
            index = self.header_row_index
        elif self.first_data_row_index:
            index = self.first_data_row_index - 1
        return index

    def parse_data(self, raise_errors: bool = False) -> None:
        try:
            self._parse()
        except Exception as e:
            messages = getattr(e, 'messages', str(e))
            self.add_errors(messages)

        if raise_errors and self.has_errors:
            raise ParserError(self.errors)

    def _load_workbook(self) -> Workbook:
        """Загрузка openpyxl Workbook из файла."""
        return load_workbook(
            filename=self.file_path or self.file_contents,
            read_only=self.read_only_workbook,
            data_only=self.read_only_workbook,
        )

    def _iterate_worksheet_rows(self, worksheet: Worksheet) -> Iterator[Tuple[int, List[Any]]]:
        row_index = self.first_data_row_index
        min_row = self.first_data_row_index + 1 if self.first_data_row_index is not None else None
        max_row = self.last_data_row_index + 1 if self.last_data_row_index is not None else None
        for row in worksheet.iter_rows(min_row=min_row, max_row=max_row):
            yield row_index, [cell.value for cell in row]
            row_index += 1

    def _validate_worksheet(self, worksheet: Worksheet) -> bool:
        errors = []
        headers_error = self._validate_worksheet_headers(worksheet)
        title_error = self._validate_worksheet_title(worksheet.title)
        if headers_error:
            errors.append(headers_error)
        if title_error:
            errors.append(title_error)
        if errors:
            self.add_errors(errors, worksheet_title=worksheet.title)
            return False
        return True

    def _validate_worksheet_headers(self, worksheet: Worksheet) -> Optional[str]:
        expected_headers = {
            column.index: column.header.lower()
            for column in self.columns
            if column.header and column.validate_header
        }
        if expected_headers and self.header_row_offset is not None:
            for row in worksheet.iter_rows(min_row=self.header_row_offset + 1):
                columns = {
                    idx: col.value.strip().lower()
                    if isinstance(col.value, str) else col.value
                    for idx, col in enumerate(row) if idx in expected_headers
                }

                if columns != expected_headers:
                    file_path = self.file_path or 'file'
                    return (
                        f'Incorrect column names in the file: {file_path}. '
                        f'Worksheet title: {worksheet.title}. '
                        f'Columns in file: {columns}. '
                        f'Expected columns: {expected_headers}.'
                    )
                break

    def _validate_worksheet_title(self, title: str) -> Optional[str]:
        """Опциональный метод для валидации заголовков страниц файла."""
        pass

    def _parse(self) -> None:
        self.workbook = self._load_workbook()
        for worksheet in self.workbook.worksheets:
            if not self._validate_worksheet(worksheet):
                continue
            self._parse_worksheet(worksheet)

    def _parse_worksheet(self, worksheet: Worksheet) -> None:
        data = []

        for row_index, row in self._iterate_worksheet_rows(worksheet):
            try:
                row_data = self.parse_row(row, row_index, worksheet_title=worksheet.title)
            except SkipRow:
                pass
            else:
                data.append(row_data)

        self.cleaned_data.extend(self.clean(data))
